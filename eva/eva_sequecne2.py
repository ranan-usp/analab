import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# Excelファイルの読み込み
file_path = 'en.xlsx'  # ここにExcelファイルのパスを入力してください
xls = pd.ExcelFile(file_path)

# データフレームに読み込み
df = pd.read_excel(xls, sheet_name=0)  # 最初のシートを読み込み

# 新しいシートに保存するデータを格納するリスト
sheets_data = []
sheet_names = []
start_index = None

# !Sequenceに一致する行を探索し、対応する範囲を抽出
for i in range(len(df)):
    if df.iloc[i, 0] == '!Sequence':
        if start_index is not None:
            # 前のシーケンスが終了する行までを切り出して保存
            sheet_data = df.iloc[start_index:i, :]
            sheets_data.append(sheet_data)
            # シート名を決定（!Sequenceの次の行のセル文字列）
            sheet_names.append(df.iloc[start_index + 1, 0])
        start_index = i
# 最後の!Sequenceからファイルの終わりまでのデータを追加
if start_index is not None:
    sheet_data = df.iloc[start_index:, :]
    sheets_data.append(sheet_data)
    sheet_names.append(df.iloc[start_index + 1, 0])

# 新しいExcelファイルに各シーケンスをシートとして保存（Arialフォント設定）
with pd.ExcelWriter('output_sequences2.xlsx', engine='openpyxl') as writer:
    for idx, (sheet_data, sheet_name) in enumerate(zip(sheets_data, sheet_names)):
        # データを新しいシートに書き込み
        wb = writer.book
        ws = wb.create_sheet(title=sheet_name)
        
        for r_idx, row in enumerate(dataframe_to_rows(sheet_data, index=False, header=True)):
            ws.append(row)
            if r_idx == 0:  # ヘッダー行の場合
                for cell in ws[r_idx + 1]:
                    cell.font = Font(name='Arial', bold=True)
            else:
                for cell in ws[r_idx + 1]:
                    cell.font = Font(name='Arial')

# 既存のシートを削除
del wb['Sheet']

# 新しいシートで保存
wb.save('output_sequences2.xlsx')
