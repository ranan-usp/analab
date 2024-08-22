import pandas as pd
from openpyxl import load_workbook
import re,pprint,sys

# Excelファイルとシート名の指定
file_path = 'output_sequences2.xlsx'  # 作成したExcelファイルのパスを指定
sheet_name = 'binary_search'  # 解析したいシート名を指定

# Excelファイルの読み込み
wb = load_workbook(filename=file_path)
ws = wb[sheet_name]

# 解析結果を格納するリスト
group_data = dict()
current_group = []
sequence_type = None

# シート内の1列目のセルを走査し、!Sequenceブロックを抽出
for row in ws.iter_rows(values_only=True):
    cell_value = row[0]
    
    if isinstance(cell_value, str) and "!Sequence" in cell_value:
        if current_group:
            group_data[sequence_type] = current_group
            current_group = []
        sequence_type = cell_value
    current_group.append(list(row))
    
if current_group:  # 最後のグループを追加
    group_data[sequence_type] = current_group

Condition_dict = dict()

InitialEvent_dict = dict()
UserFunction_dict = dict()
LoopFunction_dict = dict()
# 各グループを解析して結果を表示
for sequence_type, group in group_data.items():
    
    print(f"\nProcessing {sequence_type}...")

    data = pd.DataFrame(group[1:],columns=group[0])

    if sequence_type == "!Sequence_MVICondition":
        for index, row in data.iterrows():
            Condition_dict[row["!Sequence_MVICondition"]] = {"Output":row["Output"],"Type":"MVI","Measure":row["MeasureType"]}
    if sequence_type == "!Sequence_AVICondition":
        for index, row in data.iterrows():
            Condition_dict[row["!Sequence_AVICondition"]] = {"Output":row["Output"],"Type":"AVI","Measure":row["MeasureType"]}
    if sequence_type == "!Sequence_Resource":
        for index, row in data.iterrows():
            InitialEvent_dict[row["!Sequence_Resource"]] = {"SetupCondition":row["SetupCondition"]}

    if sequence_type == "!Sequence_UserFunction":
        
        for index, row in data.iterrows():
            print(row)
            flg = 0
            var_list = list()
            for i, v in row.items():
                if flg == 1 and v != None:
                    var_list.append(row.iloc[-1])
                if flg == 0 and i == "Args":
                    flg = 1

            UserFunction_dict[row["!Sequence_UserFunction"]] = {"Function":row["Function"],"Var":var_list}

        sys.exit()
    
    if sequence_type == "!Sequence_LoopTracking":
        for index, row in data.iterrows():
            LoopFunction_dict[row["!Sequence_LoopTracking"]] = {"LoopTracking":row["TrackVariable"] + "=" + row["Expression"]}


Event_Sequence = list()
group = group_data["!Sequence_Event"]
data = pd.DataFrame(group[1:],columns=group[0])

pprint.pprint(InitialEvent_dict)


for index, row in data.iterrows():
    if row["Type"] == "PowerOnEvent":
        Event_Sequence.append({"What":row["Resource"],
                               "Type":row["Type"],
                               "Do":Condition_dict[InitialEvent_dict[row["Resource"]]["SetupCondition"]]["Output"]})
    elif row["Type"] == "MeasureEvent":
        Event_Sequence.append({"What":row["Resource"],
                               "Type":row["Type"],
                               "Do":Condition_dict[InitialEvent_dict[row["Resource"]]["SetupCondition"]]["Measure"]})
    elif row["Type"] == "SetupEvent":
        Event_Sequence.append({"What":row["Resource"],
                               "Type":row["Type"],
                               "Do":Condition_dict[row["SetupCondition"]]["Output"]})
    elif row["Type"] == "CallFunctionEvent":
        Event_Sequence.append({"What":row["Resource"],
                               "Type":row["Type"] + "-" + UserFunction_dict[row["!Sequence_Event"]]["Function"],
                               "Do":"jfoarj"})
    elif row["Type"] == "LoopBegin":
        if row["Resource"] in LoopFunction_dict.keys():
            Event_Sequence.append({"What":row["Resource"],
                                "Type":row["Type"],
                                "Do":LoopFunction_dict[row["Resource"]]["LoopTracking"]})
        else:
            Event_Sequence.append({"What":row["Resource"],
                                "Type":row["Type"],
                                "Do":row["Type"]})
    elif row["Type"] == "LoopBreak":
        Event_Sequence.append({"What":row["Resource"],
                                "Type":row["Type"],
                                "Do":row["LoopBreakContinue"]})
    elif row["Type"] == "LoopEnd":
        Event_Sequence.append({"What":row["Resource"],
                                "Type":row["Type"],
                                "Do":row["Type"]})
    elif row["Type"] in ["If", "Else", "EndIf"]:
        Event_Sequence.append({"What":row["Resource"],
                               "Type":row["Type"],
                               "Do":row["CondtionalBranchExpression"]})
    
    elif row["Type"] in ["SetVariableEvent", "PowerOffEvent"]:
        pass


for event in Event_Sequence:
    print(event["What"],",",event["Type"],",",event["Do"])

